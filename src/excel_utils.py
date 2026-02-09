"""
Utilidades para manejo de archivos Excel.
"""
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from pathlib import Path
from typing import Optional
from logger_config import setup_logger
from constants import (
    EXCEL_HEADER_COLOR, EXCEL_FONT_COLOR, EXCEL_FONT_SIZE,
    EXCEL_BORDER_COLOR, EXCEL_COLUMN_WIDTH
)

logger = setup_logger('excel_utils')


def get_excel_styles():
    """
    Retorna los estilos estándar para Excel.
    
    Returns:
        Tupla con (borde, relleno, fuente)
    """
    borde = Border(
        left=Side(border_style='thin', color=EXCEL_BORDER_COLOR),
        right=Side(border_style='thin', color=EXCEL_BORDER_COLOR),
        top=Side(border_style='thin', color=EXCEL_BORDER_COLOR),
        bottom=Side(border_style='thin', color=EXCEL_BORDER_COLOR)
    )
    
    relleno = PatternFill(
        start_color=EXCEL_HEADER_COLOR,
        end_color=EXCEL_HEADER_COLOR,
        fill_type='solid'
    )
    
    fuente = Font(
        bold=True,
        color=EXCEL_FONT_COLOR,
        size=EXCEL_FONT_SIZE
    )
    
    return borde, relleno, fuente


def apply_header_style(sheet, row_num: int, start_col: int = 1, end_col: int = 31):
    """
    Aplica estilo de encabezado a una fila.
    
    Args:
        sheet: Hoja de Excel
        row_num: Número de fila
        start_col: Columna inicial (1-indexed)
        end_col: Columna final (1-indexed)
    """
    borde, relleno, fuente = get_excel_styles()
    
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=row_num, column=col)
        cell.border = borde
        cell.fill = relleno
        cell.font = fuente
        
        # Ajustar ancho de columna
        col_letter = openpyxl.utils.get_column_letter(col)
        sheet.column_dimensions[col_letter].width = EXCEL_COLUMN_WIDTH


def apply_formula_to_range(sheet, range_str: str, formula_template: str = None):
    """
    Aplica fórmulas a un rango de celdas.
    
    Args:
        sheet: Hoja de Excel
        range_str: Rango en formato 'A1:B10'
        formula_template: Template de fórmula (si None, usa el valor existente)
    """
    for row in sheet[range_str]:
        for cell in row:
            if cell.value:
                if formula_template:
                    cell.value = formula_template.format(row=cell.row)
                else:
                    cell.value = f"={cell.value}"


def apply_number_format(sheet, range_str: str, format_str: str):
    """
    Aplica formato de número a un rango de celdas.
    
    Args:
        sheet: Hoja de Excel
        range_str: Rango en formato 'A1:B10'
        format_str: String de formato (ej: '#,##0.00')
    """
    for row in sheet[range_str]:
        for cell in row:
            cell.number_format = format_str


def apply_date_format(sheet, range_str: str, format_str: str = 'DD-MM-YYYY'):
    """
    Aplica formato de fecha a un rango de celdas.
    
    Args:
        sheet: Hoja de Excel
        range_str: Rango en formato 'A1:B10'
        format_str: String de formato de fecha
    """
    apply_number_format(sheet, range_str, format_str)


def apply_bold_font(sheet, range_str: str):
    """
    Aplica fuente en negrita a un rango de celdas.
    
    Args:
        sheet: Hoja de Excel
        range_str: Rango en formato 'A1:B10'
    """
    for row in sheet[range_str]:
        for cell in row:
            cell.font = Font(bold=True)


def clear_range(sheet, range_str: str):
    """
    Limpia el contenido de un rango de celdas.
    
    Args:
        sheet: Hoja de Excel
        range_str: Rango en formato 'A1:B10'
    """
    for row in sheet[range_str]:
        for cell in row:
            cell.value = ''


def validate_template_exists(template_name: str) -> bool:
    """
    Valida que un archivo template exista.
    
    Args:
        template_name: Nombre del archivo template
        
    Returns:
        True si existe, False en caso contrario
    """
    template_path = Path(template_name)
    
    if not template_path.exists():
        logger.error(f"Template no encontrado: {template_name}")
        return False
    
    if not template_path.suffix == '.xlsx':
        logger.error(f"El archivo no es un Excel válido: {template_name}")
        return False
    
    logger.info(f"Template validado: {template_name}")
    return True


def set_column_widths(sheet, widths: dict):
    """
    Establece anchos de columnas específicos.
    
    Args:
        sheet: Hoja de Excel
        widths: Diccionario con {columna: ancho}, ej: {'A': 12, 'B': 20}
    """
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
